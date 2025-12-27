import csv
from datetime import datetime
import io
import os
import pandas as pd
import requests

if 'STREAMLIT_SERVER_HEADLESS' in os.environ:
    import streamlit as st
    import pandas as pd
    MODE = 'web'
else:
    try:
        import tkinter as tk
        from tkinter import ttk, messagebox, scrolledtext, filedialog
        import pandas as pd
        MODE = 'desktop'
    except ImportError:
        import streamlit as st
        import pandas as pd
        MODE = 'web'

from PIL import Image
if MODE == 'desktop':
    from PIL import ImageTk
    from reportlab.lib.pagesizes import letter
    from reportlab.pdfgen import canvas
    from openpyxl import Workbook

# GitHub Gist configuration for persistent storage
GIST_ID = "180cd598a8a46a4e554eeb6e2c9b8c0e"
TOKEN = "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"
GIST_URL = f"https://api.github.com/gists/{GIST_ID}"
HEADERS = {"Authorization": f"token {TOKEN}"}

def load_csv_from_gist():
    response = requests.get(GIST_URL)
    if response.status_code == 200:
        gist_data = response.json()
        csv_content = gist_data['files']['expenses.csv']['content']
        return csv_content
    else:
        return "التاريخ,القسم,المبلغ,ملاحظات\n"

def save_csv_to_gist(csv_content):
    data = {
        "files": {
            "expenses.csv": {
                "content": csv_content
            }
        }
    }
    response = requests.patch(GIST_URL, headers=HEADERS, json=data)
    return response.status_code == 200

def load_csv_local():
    try:
        with open(FILE_NAME, "r", encoding="utf-8") as f:
            return f.read()
    except FileNotFoundError:
        return "التاريخ,القسم,المبلغ,ملاحظات\n"

def save_csv_local(csv_content):
    with open(FILE_NAME, "w", encoding="utf-8") as f:
        f.write(csv_content)

FILE_NAME = "expenses.csv"
CATEGORIES_FILE = "categories.txt"

DEFAULT_CATEGORIES = [
    "مرتبات",
    "مدفوعات",
    "مشتريات",
    "زباله",
    "كهرباء",
    "انترنت",
    "باقه موبيل",
    "صيانه",
    "ايجار",
    "فيزا",
    "كاش",
    "مصروفات خاصة"
]

def load_categories():
    try:
        with open(CATEGORIES_FILE, "r", encoding="utf-8") as f:
            return [line.strip() for line in f if line.strip()]
    except FileNotFoundError:
        return DEFAULT_CATEGORIES.copy()

def save_categories():
    with open(CATEGORIES_FILE, "w", encoding="utf-8") as f:
        for cat in CATEGORIES:
            f.write(cat + "\n")

CATEGORIES = load_categories()

def init_file():
    try:
        with open(FILE_NAME, "x", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(["التاريخ", "القسم", "المبلغ", "ملاحظات"])
    except FileExistsError:
        pass

def add_expense():
    print("\nاختر القسم:")
    for i, cat in enumerate(CATEGORIES, 1):
        print(f"{i}. {cat}")

    choice = int(input("رقم القسم: "))
    category = CATEGORIES[choice - 1]

    amount = float(input("المبلغ: "))
    notes = input("ملاحظات (اختياري): ")
    date = datetime.now().strftime("%Y-%m-%d")

    with open(FILE_NAME, "a", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow([date, category, amount, notes])

    print("✅ تم إضافة المصروف بنجاح")

def show_expenses():
    print("\n📄 كل المصروفات:")
    with open(FILE_NAME, "r", encoding="utf-8") as f:
        reader = csv.reader(f)
        for row in reader:
            print(" | ".join(row))

def total_by_category():
    totals = {cat: 0 for cat in CATEGORIES}

    with open(FILE_NAME, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            totals[row["القسم"]] += float(row["المبلغ"])

    print("\n📊 إجمالي المصروفات حسب القسم:")
    for cat, total in totals.items():
        print(f"{cat}: {total:.2f} جنيه")

def add_expense_gui(root):
    def validate_amount(P):
        if P == "" or P == ".":
            return True
        try:
            float(P)
            return True
        except ValueError:
            return False

    def submit():
        try:
            category = category_var.get().strip()
            amount_str = amount_entry.get().strip()
            notes = notes_entry.get().strip()
            date_str = date_entry.get().strip()

            if not category:
                messagebox.showerror("خطأ", "يجب اختيار القسم")
                return
            if not amount_str:
                messagebox.showerror("خطأ", "يجب إدخال المبلغ")
                return
            amount = float(amount_str)
            if amount <= 0:
                messagebox.showerror("خطأ", "المبلغ يجب أن يكون أكبر من صفر")
                return
            if not date_str:
                messagebox.showerror("خطأ", "يجب إدخال التاريخ")
                return
            # Validate date format
            try:
                datetime.strptime(date_str, "%Y-%m-%d")
            except ValueError:
                messagebox.showerror("خطأ", "التاريخ يجب أن يكون بالصيغة YYYY-MM-DD")
                return

            with open(FILE_NAME, "a", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                writer.writerow([date_str, category, amount, notes])

            messagebox.showinfo("نجح", "تم إضافة المصروف بنجاح")
            add_window.destroy()
        except ValueError:
            messagebox.showerror("خطأ", "المبلغ يجب أن يكون رقماً صحيحاً")

    add_window = tk.Toplevel(root)
    add_window.title("إضافة مصروف")

    tk.Label(add_window, text="القسم:").grid(row=0, column=0)
    category_var = tk.StringVar()
    category_combo = ttk.Combobox(add_window, textvariable=category_var, values=CATEGORIES)
    category_combo.grid(row=0, column=1)

    tk.Label(add_window, text="المبلغ:").grid(row=1, column=0)
    vcmd = (add_window.register(validate_amount), '%P')
    amount_entry = tk.Entry(add_window, validate="key", validatecommand=vcmd)
    amount_entry.grid(row=1, column=1)

    tk.Label(add_window, text="التاريخ (YYYY-MM-DD):").grid(row=2, column=0)
    date_entry = tk.Entry(add_window)
    date_entry.insert(0, datetime.now().strftime("%Y-%m-%d"))
    date_entry.grid(row=2, column=1)

    tk.Label(add_window, text="ملاحظات:").grid(row=3, column=0)
    notes_entry = tk.Entry(add_window)
    notes_entry.grid(row=3, column=1)

    tk.Button(add_window, text="إضافة", command=submit).grid(row=4, column=0, columnspan=2)

def show_expenses_gui(root):
    def delete_expense():
        selected_item = tree.selection()
        if not selected_item:
            messagebox.showerror("خطأ", "يجب اختيار مصروف للحذف")
            return
        item = tree.item(selected_item)
        values = item['values']
        if values[0] == "التاريخ":  # Header
            return
        confirm = messagebox.askyesno("تأكيد", f"هل تريد حذف المصروف: {values[0]} - {values[1]} - {values[2]} - {values[3]}؟")
        if not confirm:
            return
        # Read all expenses
        expenses = []
        try:
            with open(FILE_NAME, "r", encoding="utf-8") as f:
                reader = csv.reader(f)
                header = next(reader)
                expenses.append(header)
                for row in reader:
                    if row != list(values):
                        expenses.append(row)
        except FileNotFoundError:
            pass
        # Write back without the deleted expense
        with open(FILE_NAME, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerows(expenses)
        # Refresh the tree
        tree.delete(*tree.get_children())
        for expense in expenses[1:]:
            tree.insert("", tk.END, values=expense)
        messagebox.showinfo("نجح", "تم حذف المصروف بنجاح")

    show_window = tk.Toplevel(root)
    show_window.title("عرض كل المصروفات")

    tree = ttk.Treeview(show_window, columns=("التاريخ", "القسم", "المبلغ", "ملاحظات"), show="headings")
    tree.heading("التاريخ", text="التاريخ")
    tree.heading("القسم", text="القسم")
    tree.heading("المبلغ", text="المبلغ")
    tree.heading("ملاحظات", text="ملاحظات")
    tree.pack(fill=tk.BOTH, expand=True)

    try:
        with open(FILE_NAME, "r", encoding="utf-8") as f:
            reader = csv.reader(f)
            header = next(reader)
            for row in reader:
                tree.insert("", tk.END, values=row)
    except FileNotFoundError:
        pass

    button_frame = tk.Frame(show_window)
    button_frame.pack(pady=10)
    tk.Button(button_frame, text="حذف المصروف المحدد", command=delete_expense).pack()

def total_by_category_gui(root):
    from collections import defaultdict
    totals = defaultdict(float)

    try:
        with open(FILE_NAME, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                totals[row["القسم"]] += float(row["المبلغ"])
    except FileNotFoundError:
        pass

    total_window = tk.Toplevel(root)
    total_window.title("إجمالي المصروفات حسب القسم")
    total_window.geometry("400x300")
    total_window.deiconify()
    total_window.focus()
    total_window.lift()
    total_window.state('normal')
    total_window.attributes("-topmost", True)

    text = scrolledtext.ScrolledText(total_window, width=50, height=20)
    text.pack()

    has_data = any(total > 0 for total in totals.values())
    if not has_data:
        text.insert(tk.END, "لا توجد مصروفات بعد")
    else:
        for cat, total in sorted(totals.items()):
            if total > 0:
                text.insert(tk.END, f"{cat}: {total:.2f} جنيه\n")

def monthly_reports_gui(root):
    monthly_totals = {}

    try:
        with open(FILE_NAME, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                date = row["التاريخ"]
                month_year = date[:7]  # YYYY-MM
                amount = float(row["المبلغ"])
                if month_year not in monthly_totals:
                    monthly_totals[month_year] = 0
                monthly_totals[month_year] += amount
    except FileNotFoundError:
        pass

    def search():
        year = year_entry.get().strip()
        month = month_entry.get().strip()
        if not year or not month:
            messagebox.showerror("خطأ", "يجب إدخال السنة والشهر")
            return
        try:
            month_int = int(month)
            if month_int < 1 or month_int > 12:
                raise ValueError
            year_int = int(year)
        except ValueError:
            messagebox.showerror("خطأ", "السنة والشهر يجب أن يكونا أرقام صحيحة")
            return
        month_year = f"{year}-{month.zfill(2)}"
        if month_year in monthly_totals:
            text.delete(1.0, tk.END)
            text.insert(tk.END, f"{month_year}: {monthly_totals[month_year]:.2f} جنيه\n")
        else:
            text.delete(1.0, tk.END)
            text.insert(tk.END, "لا توجد مصروفات لهذا الشهر")

    def show_all():
        text.delete(1.0, tk.END)
        if monthly_totals:
            for month, total in sorted(monthly_totals.items()):
                text.insert(tk.END, f"{month}: {total:.2f} جنيه\n")
        else:
            text.insert(tk.END, "لا توجد مصروفات بعد")

    def export_pdf():
        file_path = filedialog.asksaveasfilename(defaultextension=".pdf", filetypes=[("PDF files", "*.pdf")])
        if file_path:
            c = canvas.Canvas(file_path, pagesize=letter)
            c.drawString(100, 750, "تقارير شهرية")
            y = 720
            content = text.get(1.0, tk.END).strip()
            if content and not content.startswith("لا توجد"):
                lines = content.split('\n')
                for line in lines:
                    if line:
                        c.drawString(100, y, line)
                        y -= 20
            else:
                for month, total in sorted(monthly_totals.items()):
                    c.drawString(100, y, f"{month}: {total:.2f} جنيه")
                    y -= 20
            c.save()
            messagebox.showinfo("نجح", "تم تصدير التقرير إلى PDF")

    def export_excel():
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if file_path:
            wb = Workbook()
            ws = wb.active
            ws.title = "تقارير شهرية"
            ws['A1'] = "الشهر"
            ws['B1'] = "المجموع"
            row = 2
            content = text.get(1.0, tk.END).strip()
            if content and not content.startswith("لا توجد"):
                lines = content.split('\n')
                for line in lines:
                    if line and ': ' in line:
                        month, total_str = line.split(': ')
                        total = total_str.replace(' جنيه', '')
                        ws[f'A{row}'] = month
                        ws[f'B{row}'] = float(total)
                        row += 1
            else:
                for month, total in sorted(monthly_totals.items()):
                    ws[f'A{row}'] = month
                    ws[f'B{row}'] = total
                    row += 1
            wb.save(file_path)
            messagebox.showinfo("نجح", "تم تصدير التقرير إلى Excel")

    report_window = tk.Toplevel(root)
    report_window.title("تقارير شهرية")

    search_frame = tk.Frame(report_window)
    search_frame.pack(pady=10)

    tk.Label(search_frame, text="السنة:").grid(row=0, column=0)
    year_entry = tk.Entry(search_frame)
    year_entry.grid(row=0, column=1)

    tk.Label(search_frame, text="الشهر (1-12):").grid(row=0, column=2)
    month_entry = tk.Entry(search_frame)
    month_entry.grid(row=0, column=3)

    tk.Button(search_frame, text="بحث", command=search).grid(row=0, column=4, padx=5)
    tk.Button(search_frame, text="عرض الكل", command=show_all).grid(row=0, column=5, padx=5)

    text = scrolledtext.ScrolledText(report_window, width=50, height=20)
    text.pack()

    button_frame = tk.Frame(report_window)
    button_frame.pack(pady=10)

    tk.Button(button_frame, text="تصدير PDF", command=export_pdf).pack(side=tk.LEFT, padx=5)
    tk.Button(button_frame, text="تصدير Excel", command=export_excel).pack(side=tk.LEFT, padx=5)

    show_all()

def detailed_monthly_reports_gui(root):
    monthly_expenses = {}

    try:
        with open(FILE_NAME, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                date = row["التاريخ"]
                month_year = date[:7]  # YYYY-MM
                if month_year not in monthly_expenses:
                    monthly_expenses[month_year] = []
                monthly_expenses[month_year].append(row)
    except FileNotFoundError:
        pass

    def export_pdf():
        file_path = filedialog.asksaveasfilename(defaultextension=".pdf", filetypes=[("PDF files", "*.pdf")])
        if file_path:
            c = canvas.Canvas(file_path, pagesize=letter)
            c.drawString(100, 750, "تقارير شهرية مفصلة")
            y = 720
            for month, expenses in sorted(monthly_expenses.items()):
                c.drawString(100, y, f"الشهر: {month}")
                y -= 20
                for expense in expenses:
                    c.drawString(120, y, f"التاريخ: {expense['التاريخ']}, القسم: {expense['القسم']}, المبلغ: {expense['المبلغ']}, ملاحظات: {expense['ملاحظات']}")
                    y -= 20
                    if y < 50:
                        c.showPage()
                        y = 750
                total = sum(float(expense['المبلغ']) for expense in expenses)
                c.drawString(100, y, f"الإجمالي: {total:.2f} جنيه")
                y -= 30
                if y < 50:
                    c.showPage()
                    y = 750
            c.save()
            messagebox.showinfo("نجح", "تم تصدير التقرير إلى PDF")

    def export_excel():
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if file_path:
            wb = Workbook()
            ws = wb.active
            ws.title = "تقارير شهرية مفصلة"
            ws['A1'] = "الشهر"
            ws['B1'] = "التاريخ"
            ws['C1'] = "القسم"
            ws['D1'] = "المبلغ"
            ws['E1'] = "ملاحظات"
            row = 2
            for month, expenses in sorted(monthly_expenses.items()):
                for expense in expenses:
                    ws[f'A{row}'] = month
                    ws[f'B{row}'] = expense['التاريخ']
                    ws[f'C{row}'] = expense['القسم']
                    ws[f'D{row}'] = expense['المبلغ']
                    ws[f'E{row}'] = expense['ملاحظات']
                    row += 1
                total = sum(float(expense['المبلغ']) for expense in expenses)
                ws[f'A{row}'] = month
                ws[f'B{row}'] = "الإجمالي"
                ws[f'D{row}'] = total
                row += 1
            wb.save(file_path)
            messagebox.showinfo("نجح", "تم تصدير التقرير إلى Excel")

    report_window = tk.Toplevel(root)
    report_window.title("تقارير شهرية مفصلة")

    text = scrolledtext.ScrolledText(report_window, width=80, height=30)
    text.pack()

    button_frame = tk.Frame(report_window)
    button_frame.pack(pady=10)

    tk.Button(button_frame, text="تصدير PDF", command=export_pdf).pack(side=tk.LEFT, padx=5)
    tk.Button(button_frame, text="تصدير Excel", command=export_excel).pack(side=tk.LEFT, padx=5)

    if monthly_expenses:
        for month, expenses in sorted(monthly_expenses.items()):
            total = sum(float(expense['المبلغ']) for expense in expenses)
            text.insert(tk.END, f"الشهر: {month}\n")
            for expense in expenses:
                text.insert(tk.END, f"  التاريخ: {expense['التاريخ']}, القسم: {expense['القسم']}, المبلغ: {expense['المبلغ']}, ملاحظات: {expense['ملاحظات']}\n")
            text.insert(tk.END, f"الإجمالي: {total:.2f} جنيه\n\n")
    else:
        text.insert(tk.END, "لا توجد مصروفات بعد")

def add_category_gui(root):
    def submit():
        new_cat = cat_entry.get().strip()
        if new_cat and new_cat not in CATEGORIES:
            CATEGORIES.append(new_cat)
            save_categories()
            messagebox.showinfo("نجح", f"تم إضافة القسم: {new_cat}")
            add_cat_window.destroy()
        elif new_cat in CATEGORIES:
            messagebox.showerror("خطأ", "القسم موجود بالفعل")
        else:
            messagebox.showerror("خطأ", "أدخل اسم القسم")

    def delete_cat():
        cat_to_delete = delete_var.get().strip()
        if not cat_to_delete:
            messagebox.showerror("خطأ", "يجب اختيار القسم للحذف")
            return
        if cat_to_delete in DEFAULT_CATEGORIES:
            messagebox.showerror("خطأ", "لا يمكن حذف الأقسام الافتراضية")
            return
        # Check if category is used in expenses
        try:
            with open(FILE_NAME, "r", encoding="utf-8") as f:
                reader = csv.DictReader(f)
                for row in reader:
                    if row["القسم"] == cat_to_delete:
                        messagebox.showerror("خطأ", "لا يمكن حذف القسم لأنه مستخدم في مصروفات")
                        return
        except FileNotFoundError:
            pass
        # Remove from CATEGORIES
        CATEGORIES.remove(cat_to_delete)
        save_categories()
        messagebox.showinfo("نجح", f"تم حذف القسم: {cat_to_delete}")
        add_cat_window.destroy()

    add_cat_window = tk.Toplevel(root)
    add_cat_window.title("إدارة الأقسام")

    tk.Label(add_cat_window, text="اسم القسم الجديد:").grid(row=0, column=0)
    cat_entry = tk.Entry(add_cat_window)
    cat_entry.grid(row=0, column=1)

    tk.Button(add_cat_window, text="إضافة", command=submit).grid(row=1, column=0, columnspan=2)

    tk.Label(add_cat_window, text="حذف قسم موجود:").grid(row=2, column=0)
    delete_var = tk.StringVar()
    delete_combo = ttk.Combobox(add_cat_window, textvariable=delete_var, values=CATEGORIES)
    delete_combo.grid(row=2, column=1)

    tk.Button(add_cat_window, text="حذف", command=delete_cat).grid(row=3, column=0, columnspan=2)

def reports_and_closure_gui(root):
    def validate_amount(P):
        if P == "" or P == ".":
            return True
        try:
            float(P)
            return True
        except ValueError:
            return False

    def daily_closure_submit():
        try:
            date_str = date_entry.get().strip()
            visa_str = visa_entry.get().strip()
            cash_str = cash_entry.get().strip()
            expenses_str = expenses_entry.get().strip()
            notes = notes_entry.get().strip()

            if not date_str:
                messagebox.showerror("خطأ", "يجب إدخال التاريخ")
                return
            # Validate date format
            try:
                datetime.strptime(date_str, "%Y-%m-%d")
            except ValueError:
                messagebox.showerror("خطأ", "التاريخ يجب أن يكون بالصيغة YYYY-MM-DD")
                return

            visa_amount = float(visa_str) if visa_str else 0
            cash_amount = float(cash_str) if cash_str else 0
            expenses_amount = float(expenses_str) if expenses_str else 0

            if visa_amount < 0 or cash_amount < 0 or expenses_amount < 0:
                messagebox.showerror("خطأ", "المبالغ يجب أن تكون موجبة أو صفر")
                return

            with open(FILE_NAME, "a", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                if visa_amount > 0:
                    writer.writerow([date_str, "فيزا", visa_amount, notes])
                if cash_amount > 0:
                    writer.writerow([date_str, "كاش", cash_amount, notes])
                if expenses_amount > 0:
                    writer.writerow([date_str, "مصروفات", expenses_amount, notes])

            messagebox.showinfo("نجح", "تم حفظ الإدخال بنجاح")
            # Clear fields for new entry
            visa_entry.delete(0, tk.END)
            cash_entry.delete(0, tk.END)
            expenses_entry.delete(0, tk.END)
            notes_entry.delete(0, tk.END)
        except ValueError:
            messagebox.showerror("خطأ", "المبالغ يجب أن تكون أرقام صحيحة")

    def delete_daily_entries():
        date_str = date_entry.get().strip()
        if not date_str:
            messagebox.showerror("خطأ", "يجب إدخال التاريخ")
            return
        # Validate date format
        try:
            datetime.strptime(date_str, "%Y-%m-%d")
        except ValueError:
            messagebox.showerror("خطأ", "التاريخ يجب أن يكون بالصيغة YYYY-MM-DD")
            return

        confirm = messagebox.askyesno("تأكيد", f"هل تريد حذف إدخالات إغلاق اليوم للتاريخ {date_str}؟")
        if not confirm:
            return

        # Read all expenses
        expenses = []
        try:
            with open(FILE_NAME, "r", encoding="utf-8") as f:
                reader = csv.reader(f)
                header = next(reader)
                expenses.append(header)
                for row in reader:
                    if not (row[0] == date_str and row[1] in ["فيزا", "كاش", "مصروفات"]):
                        expenses.append(row)
        except FileNotFoundError:
            pass
        # Write back without the deleted entries
        with open(FILE_NAME, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerows(expenses)
        messagebox.showinfo("نجح", f"تم حذف إدخالات إغلاق اليوم للتاريخ {date_str}")

    def generate_daily_report():
        today = datetime.now().strftime("%Y-%m-%d")
        daily_expenses = []
        total = 0
        try:
            with open(FILE_NAME, "r", encoding="utf-8") as f:
                reader = csv.DictReader(f)
                for row in reader:
                    if row["التاريخ"] == today:
                        daily_expenses.append(row)
                        total += float(row["المبلغ"])
        except FileNotFoundError:
            pass

        text.delete(1.0, tk.END)
        if daily_expenses:
            text.insert(tk.END, f"تقرير اليوم: {today}\n\n")
            for expense in daily_expenses:
                text.insert(tk.END, f"القسم: {expense['القسم']}, المبلغ: {expense['المبلغ']}, ملاحظات: {expense['ملاحظات']}\n")
            text.insert(tk.END, f"\nالإجمالي: {total:.2f} جنيه")
        else:
            text.insert(tk.END, f"لا توجد مصروفات لليوم {today}")

    def export_daily_excel():
        today = datetime.now().strftime("%Y-%m-%d")
        daily_expenses = []
        total = 0
        try:
            with open(FILE_NAME, "r", encoding="utf-8") as f:
                reader = csv.DictReader(f)
                for row in reader:
                    if row["التاريخ"] == today:
                        daily_expenses.append(row)
                        total += float(row["المبلغ"])
        except FileNotFoundError:
            pass

        if not daily_expenses:
            messagebox.showerror("خطأ", "لا توجد مصروفات للتصدير")
            return

        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if file_path:
            wb = Workbook()
            ws = wb.active
            ws.title = f"تقرير يومي {today}"
            ws['A1'] = "التاريخ"
            ws['B1'] = "القسم"
            ws['C1'] = "المبلغ"
            ws['D1'] = "ملاحظات"
            row = 2
            for expense in daily_expenses:
                ws[f'A{row}'] = expense['التاريخ']
                ws[f'B{row}'] = expense['القسم']
                ws[f'C{row}'] = expense['المبلغ']
                ws[f'D{row}'] = expense['ملاحظات']
                row += 1
            ws[f'A{row}'] = "الإجمالي"
            ws[f'C{row}'] = total
            wb.save(file_path)
            messagebox.showinfo("نجح", "تم تصدير التقرير اليومي إلى Excel")

    def generate_monthly_report():
        current_month = datetime.now().strftime("%Y-%m")
        monthly_expenses = []
        total = 0
        try:
            with open(FILE_NAME, "r", encoding="utf-8") as f:
                reader = csv.DictReader(f)
                for row in reader:
                    if row["التاريخ"].startswith(current_month):
                        monthly_expenses.append(row)
                        total += float(row["المبلغ"])
        except FileNotFoundError:
            pass

        monthly_text.delete(1.0, tk.END)
        if monthly_expenses:
            monthly_text.insert(tk.END, f"تقرير الشهر: {current_month}\n\n")
            for expense in monthly_expenses:
                monthly_text.insert(tk.END, f"التاريخ: {expense['التاريخ']}, القسم: {expense['القسم']}, المبلغ: {expense['المبلغ']}, ملاحظات: {expense['ملاحظات']}\n")
            monthly_text.insert(tk.END, f"\nالإجمالي: {total:.2f} جنيه")
        else:
            monthly_text.insert(tk.END, f"لا توجد مصروفات للشهر {current_month}")

    def export_monthly_excel():
        current_month = datetime.now().strftime("%Y-%m")
        monthly_expenses = []
        total = 0
        try:
            with open(FILE_NAME, "r", encoding="utf-8") as f:
                reader = csv.DictReader(f)
                for row in reader:
                    if row["التاريخ"].startswith(current_month):
                        monthly_expenses.append(row)
                        total += float(row["المبلغ"])
        except FileNotFoundError:
            pass

        if not monthly_expenses:
            messagebox.showerror("خطأ", "لا توجد مصروفات للتصدير")
            return

        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if file_path:
            wb = Workbook()
            ws = wb.active
            ws.title = f"تقرير شهري {current_month}"
            ws['A1'] = "التاريخ"
            ws['B1'] = "القسم"
            ws['C1'] = "المبلغ"
            ws['D1'] = "ملاحظات"
            row = 2
            for expense in monthly_expenses:
                ws[f'A{row}'] = expense['التاريخ']
                ws[f'B{row}'] = expense['القسم']
                ws[f'C{row}'] = expense['المبلغ']
                ws[f'D{row}'] = expense['ملاحظات']
                row += 1
            ws[f'A{row}'] = "الإجمالي"
            ws[f'C{row}'] = total
            wb.save(file_path)
            messagebox.showinfo("نجح", "تم تصدير التقرير الشهري إلى Excel")

    def generate_visa_cash_monthly_report():
        current_month = datetime.now().strftime("%Y-%m")
        visa_total = 0
        cash_total = 0
        special_total = 0
        expenses_total = 0
        visa_expenses = []
        cash_expenses = []
        special_expenses = []
        expenses_expenses = []
        try:
            with open(FILE_NAME, "r", encoding="utf-8") as f:
                reader = csv.DictReader(f)
                for row in reader:
                    if row["التاريخ"].startswith(current_month):
                        if row["القسم"] == "فيزا":
                            visa_total += float(row["المبلغ"])
                            visa_expenses.append(row)
                        elif row["القسم"] == "كاش":
                            cash_total += float(row["المبلغ"])
                            cash_expenses.append(row)
                        elif row["القسم"] == "مصروفات خاصة":
                            special_total += float(row["المبلغ"])
                            special_expenses.append(row)
                        elif row["القسم"] == "مصروفات":
                            expenses_total += float(row["المبلغ"])
                            expenses_expenses.append(row)
        except FileNotFoundError:
            pass

        visa_cash_text.delete(1.0, tk.END)
        if visa_expenses or cash_expenses or special_expenses or expenses_expenses:
            visa_cash_text.insert(tk.END, f"تقرير فيزا وكاش ومصروفات للشهر: {current_month}\n\n")
            if visa_expenses:
                visa_cash_text.insert(tk.END, "فيزا:\n")
                for expense in visa_expenses:
                    visa_cash_text.insert(tk.END, f"  التاريخ: {expense['التاريخ']}, المبلغ: {expense['المبلغ']}, ملاحظات: {expense['ملاحظات']}\n")
                visa_cash_text.insert(tk.END, f"إجمالي الفيزا: {visa_total:.2f} جنيه\n\n")
            else:
                visa_cash_text.insert(tk.END, "لا توجد مصروفات فيزا لهذا الشهر\n\n")
            if cash_expenses:
                visa_cash_text.insert(tk.END, "كاش:\n")
                for expense in cash_expenses:
                    visa_cash_text.insert(tk.END, f"  التاريخ: {expense['التاريخ']}, المبلغ: {expense['المبلغ']}, ملاحظات: {expense['ملاحظات']}\n")
                visa_cash_text.insert(tk.END, f"إجمالي الكاش: {cash_total:.2f} جنيه\n\n")
            else:
                visa_cash_text.insert(tk.END, "لا توجد مصروفات كاش لهذا الشهر\n\n")
            if expenses_expenses:
                visa_cash_text.insert(tk.END, "مصروفات:\n")
                for expense in expenses_expenses:
                    visa_cash_text.insert(tk.END, f"  التاريخ: {expense['التاريخ']}, المبلغ: {expense['المبلغ']}, ملاحظات: {expense['ملاحظات']}\n")
                visa_cash_text.insert(tk.END, f"إجمالي المصروفات: {expenses_total:.2f} جنيه\n\n")
            else:
                visa_cash_text.insert(tk.END, "لا توجد مصروفات لهذا الشهر\n\n")
            net_visa_cash = visa_total + cash_total - expenses_total
            visa_cash_text.insert(tk.END, f"صافي الفيزا والكاش (بعد خصم المصروفات): {net_visa_cash:.2f} جنيه\n\n")
            if special_expenses:
                visa_cash_text.insert(tk.END, "مصروفات خاصة:\n")
                for expense in special_expenses:
                    visa_cash_text.insert(tk.END, f"  التاريخ: {expense['التاريخ']}, المبلغ: {expense['المبلغ']}, ملاحظات: {expense['ملاحظات']}\n")
                visa_cash_text.insert(tk.END, f"إجمالي المصروفات الخاصة: {special_total:.2f} جنيه\n\n")
            else:
                visa_cash_text.insert(tk.END, "لا توجد مصروفات خاصة لهذا الشهر\n\n")
            total = net_visa_cash + special_total
            visa_cash_text.insert(tk.END, f"الإجمالي الكلي: {total:.2f} جنيه")
        else:
            visa_cash_text.insert(tk.END, f"لا توجد مصروفات فيزا أو كاش أو خاصة أو مصروفات للشهر {current_month}")

    def export_visa_cash_monthly_excel():
        current_month = datetime.now().strftime("%Y-%m")
        visa_total = 0
        cash_total = 0
        special_total = 0
        expenses_total = 0
        visa_expenses = []
        cash_expenses = []
        special_expenses = []
        expenses_expenses = []
        try:
            with open(FILE_NAME, "r", encoding="utf-8") as f:
                reader = csv.DictReader(f)
                for row in reader:
                    if row["التاريخ"].startswith(current_month):
                        if row["القسم"] == "فيزا":
                            visa_total += float(row["المبلغ"])
                            visa_expenses.append(row)
                        elif row["القسم"] == "كاش":
                            cash_total += float(row["المبلغ"])
                            cash_expenses.append(row)
                        elif row["القسم"] == "مصروفات خاصة":
                            special_total += float(row["المبلغ"])
                            special_expenses.append(row)
                        elif row["القسم"] == "مصروفات":
                            expenses_total += float(row["المبلغ"])
                            expenses_expenses.append(row)
        except FileNotFoundError:
            pass

        if not visa_expenses and not cash_expenses and not special_expenses and not expenses_expenses:
            messagebox.showerror("خطأ", "لا توجد مصروفات فيزا أو كاش أو خاصة أو مصروفات للتصدير")
            return

        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if file_path:
            wb = Workbook()
            ws = wb.active
            ws.title = f"تقرير فيزا وكاش ومصروفات {current_month}"
            ws['A1'] = "النوع"
            ws['B1'] = "التاريخ"
            ws['C1'] = "المبلغ"
            ws['D1'] = "ملاحظات"
            row = 2
            for expense in visa_expenses:
                ws[f'A{row}'] = "فيزا"
                ws[f'B{row}'] = expense['التاريخ']
                ws[f'C{row}'] = expense['المبلغ']
                ws[f'D{row}'] = expense['ملاحظات']
                row += 1
            for expense in cash_expenses:
                ws[f'A{row}'] = "كاش"
                ws[f'B{row}'] = expense['التاريخ']
                ws[f'C{row}'] = expense['المبلغ']
                ws[f'D{row}'] = expense['ملاحظات']
                row += 1
            for expense in special_expenses:
                ws[f'A{row}'] = "مصروفات خاصة"
                ws[f'B{row}'] = expense['التاريخ']
                ws[f'C{row}'] = expense['المبلغ']
                ws[f'D{row}'] = expense['ملاحظات']
                row += 1
            for expense in expenses_expenses:
                ws[f'A{row}'] = "مصروفات"
                ws[f'B{row}'] = expense['التاريخ']
                ws[f'C{row}'] = expense['المبلغ']
                ws[f'D{row}'] = expense['ملاحظات']
                row += 1
            row += 1
            ws[f'A{row}'] = "إجمالي الفيزا"
            ws[f'C{row}'] = visa_total
            row += 1
            ws[f'A{row}'] = "إجمالي الكاش"
            ws[f'C{row}'] = cash_total
            row += 1
            ws[f'A{row}'] = "إجمالي المصروفات"
            ws[f'C{row}'] = expenses_total
            row += 1
            net_visa_cash = visa_total + cash_total - expenses_total
            ws[f'A{row}'] = "صافي الفيزا والكاش (بعد خصم المصروفات)"
            ws[f'C{row}'] = net_visa_cash
            row += 1
            ws[f'A{row}'] = "إجمالي المصروفات الخاصة"
            ws[f'C{row}'] = special_total
            row += 1
            ws[f'A{row}'] = "الإجمالي الكلي"
            ws[f'C{row}'] = net_visa_cash + special_total
            wb.save(file_path)
            messagebox.showinfo("نجح", "تم تصدير تقرير فيزا وكاش ومصروفات الشهري إلى Excel")

    window = tk.Toplevel(root)
    window.title("تقارير وإغلاق")
    window.geometry("800x600")

    notebook = ttk.Notebook(window)
    notebook.pack(fill=tk.BOTH, expand=True)

    # Tab 1: إغلاق اليوم
    closure_frame = tk.Frame(notebook)
    notebook.add(closure_frame, text="إغلاق اليوم")

    tk.Label(closure_frame, text="التاريخ (YYYY-MM-DD):").grid(row=0, column=0, padx=10, pady=5)
    date_entry = tk.Entry(closure_frame)
    date_entry.insert(0, datetime.now().strftime("%Y-%m-%d"))
    date_entry.grid(row=0, column=1, padx=10, pady=5)

    tk.Label(closure_frame, text="مبلغ الفيزا:").grid(row=1, column=0, padx=10, pady=5)
    vcmd = (closure_frame.register(validate_amount), '%P')
    visa_entry = tk.Entry(closure_frame, validate="key", validatecommand=vcmd)
    visa_entry.grid(row=1, column=1, padx=10, pady=5)

    tk.Label(closure_frame, text="مبلغ الكاش:").grid(row=2, column=0, padx=10, pady=5)
    cash_entry = tk.Entry(closure_frame, validate="key", validatecommand=vcmd)
    cash_entry.grid(row=2, column=1, padx=10, pady=5)

    tk.Label(closure_frame, text="مبلغ المصروفات:").grid(row=3, column=0, padx=10, pady=5)
    expenses_entry = tk.Entry(closure_frame, validate="key", validatecommand=vcmd)
    expenses_entry.grid(row=3, column=1, padx=10, pady=5)

    tk.Label(closure_frame, text="ملاحظات:").grid(row=4, column=0, padx=10, pady=5)
    notes_entry = tk.Entry(closure_frame)
    notes_entry.grid(row=4, column=1, padx=10, pady=5)

    tk.Button(closure_frame, text="إدخال", command=daily_closure_submit).grid(row=5, column=0, pady=10)
    tk.Button(closure_frame, text="حذف إدخالات اليوم", command=delete_daily_entries).grid(row=5, column=1, pady=10)

    # Tab 2: تقرير يومي
    daily_frame = tk.Frame(notebook)
    notebook.add(daily_frame, text="تقرير يومي")

    tk.Button(daily_frame, text="عرض التقرير اليومي", command=generate_daily_report).pack(pady=10)
    text = scrolledtext.ScrolledText(daily_frame, width=80, height=20)
    text.pack(pady=10)
    tk.Button(daily_frame, text="تصدير إلى Excel", command=export_daily_excel).pack(pady=10)

    # Tab 3: تقرير شهري
    monthly_frame = tk.Frame(notebook)
    notebook.add(monthly_frame, text="تقرير شهري")

    tk.Button(monthly_frame, text="عرض التقرير الشهري", command=generate_monthly_report).pack(pady=10)
    monthly_text = scrolledtext.ScrolledText(monthly_frame, width=80, height=20)
    monthly_text.pack(pady=10)
    tk.Button(monthly_frame, text="تصدير إلى Excel", command=export_monthly_excel).pack(pady=10)

    # Tab 4: تقرير فيزا وكاش شهري
    visa_cash_frame = tk.Frame(notebook)
    notebook.add(visa_cash_frame, text="تقرير فيزا وكاش شهري")

    tk.Button(visa_cash_frame, text="عرض تقرير فيزا وكاش الشهري", command=generate_visa_cash_monthly_report).pack(pady=10)
    visa_cash_text = scrolledtext.ScrolledText(visa_cash_frame, width=80, height=20)
    visa_cash_text.pack(pady=10)
    tk.Button(visa_cash_frame, text="تصدير إلى Excel", command=export_visa_cash_monthly_excel).pack(pady=10)


# Web versions of the functions
def add_expense_web():
    st.header("إضافة مصروف")
    with st.form("add_expense_form"):
        category = st.selectbox("القسم", CATEGORIES)
        amount = st.number_input("المبلغ", min_value=0.0, step=0.01)
        date = st.date_input("التاريخ", value=datetime.now().date())
        notes = st.text_input("ملاحظات")
        submitted = st.form_submit_button("إضافة")
        if submitted:
            if amount <= 0:
                st.error("المبلغ يجب أن يكون أكبر من صفر")
            else:
                date_str = date.strftime("%Y-%m-%d")
                csv_content = load_csv_from_gist()
                if not csv_content.strip():
                    csv_content = "التاريخ,القسم,المبلغ,ملاحظات\n"
                csv_content += f"{date_str},{category},{amount},{notes}\n"
                if save_csv_to_gist(csv_content):
                    st.success("تم إضافة المصروف بنجاح")
                else:
                    st.error("فشل في حفظ البيانات")

def show_expenses_web():
    st.header("عرض كل المصروفات")
    csv_content = load_csv_from_gist()
    if not csv_content.strip():
        st.write("لا توجد مصروفات بعد")
    else:
        df = pd.read_csv(io.StringIO(csv_content), encoding="utf-8")
        st.dataframe(df)
        # Delete functionality
        if not df.empty:
            selected_index = st.selectbox("اختر مصروف للحذف", df.index, format_func=lambda x: f"{df.loc[x, 'التاريخ']} - {df.loc[x, 'القسم']} - {df.loc[x, 'المبلغ']} - {df.loc[x, 'ملاحظات']}")
            if st.button("حذف المصروف المحدد"):
                if st.button("تأكيد الحذف"):
                    df = df.drop(selected_index)
                    csv_content = df.to_csv(index=False)
                    if save_csv_to_gist(csv_content):
                        st.success("تم حذف المصروف بنجاح")
                        st.rerun()
                    else:
                        st.error("فشل في حفظ البيانات")

def total_by_category_web():
    st.header("إجمالي المصروفات حسب القسم")
    from collections import defaultdict
    totals = defaultdict(float)
    csv_content = load_csv_from_gist()
    if csv_content.strip():
        df = pd.read_csv(io.StringIO(csv_content), encoding="utf-8")
        for _, row in df.iterrows():
            totals[row["القسم"]] += float(row["المبلغ"])
    if totals:
        for cat, total in sorted(totals.items()):
            if total > 0:
                st.write(f"{cat}: {total:.2f} جنيه")
    else:
        st.write("لا توجد مصروفات بعد")

def monthly_reports_web():
    st.header("تقارير شهرية")
    monthly_totals = {}
    csv_content = load_csv_from_gist()
    if csv_content.strip():
        df = pd.read_csv(io.StringIO(csv_content), encoding="utf-8")
        for _, row in df.iterrows():
            date = row["التاريخ"]
            month_year = date[:7]
            amount = float(row["المبلغ"])
            if month_year not in monthly_totals:
                monthly_totals[month_year] = 0
            monthly_totals[month_year] += amount
    if monthly_totals:
        for month, total in sorted(monthly_totals.items()):
            st.write(f"{month}: {total:.2f} جنيه")
    else:
        st.write("لا توجد مصروفات بعد")

def detailed_monthly_reports_web():
    st.header("تقارير شهرية مفصلة")
    monthly_expenses = {}
    csv_content = load_csv_from_gist()
    if csv_content.strip():
        df = pd.read_csv(io.StringIO(csv_content), encoding="utf-8")
        for _, row in df.iterrows():
            date = row["التاريخ"]
            month_year = date[:7]
            if month_year not in monthly_expenses:
                monthly_expenses[month_year] = []
            monthly_expenses[month_year].append(row)
    if monthly_expenses:
        for month, expenses in sorted(monthly_expenses.items()):
            st.subheader(f"الشهر: {month}")
            total = sum(float(expense['المبلغ']) for expense in expenses)
            for expense in expenses:
                st.write(f"التاريخ: {expense['التاريخ']}, القسم: {expense['القسم']}, المبلغ: {expense['المبلغ']}, ملاحظات: {expense['ملاحظات']}")
            st.write(f"الإجمالي: {total:.2f} جنيه")
    else:
        st.write("لا توجد مصروفات بعد")

def add_category_web():
    st.header("إدارة الأقسام")
    new_cat = st.text_input("اسم القسم الجديد")
    if st.button("إضافة قسم"):
        if new_cat and new_cat not in CATEGORIES:
            CATEGORIES.append(new_cat)
            save_categories()
            st.success(f"تم إضافة القسم: {new_cat}")
        elif new_cat in CATEGORIES:
            st.error("القسم موجود بالفعل")
        else:
            st.error("أدخل اسم القسم")
    delete_cat = st.selectbox("حذف قسم موجود", CATEGORIES)
    if st.button("حذف قسم"):
        if delete_cat in DEFAULT_CATEGORIES:
            st.error("لا يمكن حذف الأقسام الافتراضية")
        else:
            # Check if category is used
            csv_content = load_csv_from_gist()
            if csv_content.strip():
                df = pd.read_csv(io.StringIO(csv_content), encoding="utf-8")
                if delete_cat in df["القسم"].values:
                    st.error("لا يمكن حذف القسم لأنه مستخدم في مصروفات")
                    return
            CATEGORIES.remove(delete_cat)
            save_categories()
            st.success(f"تم حذف القسم: {delete_cat}")

def sync_data_web():
    st.header("مزامنة البيانات")
    st.write("يمكنك هنا مزامنة البيانات بين التخزين المحلي والـ Gist.")

    col1, col2 = st.columns(2)

    with col1:
        if st.button("تحميل البيانات من Gist إلى المحلي"):
            csv_content = load_csv_from_gist()
            if csv_content.strip():
                save_csv_local(csv_content)
                st.success("تم تحميل البيانات من Gist إلى الملف المحلي بنجاح")
            else:
                st.warning("لا توجد بيانات في Gist")

    with col2:
        if st.button("رفع البيانات من المحلي إلى Gist"):
            csv_content = load_csv_local()
            if csv_content.strip():
                if save_csv_to_gist(csv_content):
                    st.success("تم رفع البيانات من الملف المحلي إلى Gist بنجاح")
                else:
                    st.error("فشل في رفع البيانات إلى Gist")
            else:
                st.warning("لا توجد بيانات في الملف المحلي")

    st.subheader("حالة البيانات")
    gist_content = load_csv_from_gist()
    local_content = load_csv_local()

    gist_lines = len(gist_content.strip().split('\n')) if gist_content.strip() else 0
    local_lines = len(local_content.strip().split('\n')) if local_content.strip() else 0

    st.write(f"عدد السطور في Gist: {gist_lines}")
    st.write(f"عدد السطور في الملف المحلي: {local_lines}")

def reports_and_closure_web():
    st.header("تقارير وإغلاق")
    tab1, tab2, tab3, tab4 = st.tabs(["إغلاق اليوم", "تقرير يومي", "تقرير شهري", "تقرير فيزا وكاش شهري"])
    with tab1:
        st.subheader("إغلاق اليوم")
        with st.form("daily_closure_form"):
            date = st.date_input("التاريخ", value=datetime.now().date())
            visa = st.number_input("مبلغ الفيزا", min_value=0.0, step=0.01)
            cash = st.number_input("مبلغ الكاش", min_value=0.0, step=0.01)
            expenses = st.number_input("مبلغ المصروفات", min_value=0.0, step=0.01)
            notes = st.text_input("ملاحظات")
            submitted = st.form_submit_button("إدخال")
            if submitted:
                date_str = date.strftime("%Y-%m-%d")
                csv_content = load_csv_from_gist()
                if not csv_content.strip():
                    csv_content = "التاريخ,القسم,المبلغ,ملاحظات\n"
                if visa > 0:
                    csv_content += f"{date_str},فيزا,{visa},{notes}\n"
                if cash > 0:
                    csv_content += f"{date_str},كاش,{cash},{notes}\n"
                if expenses > 0:
                    csv_content += f"{date_str},مصروفات,{expenses},{notes}\n"
                if save_csv_to_gist(csv_content):
                    st.success("تم حفظ الإدخال بنجاح")
                else:
                    st.error("فشل في حفظ البيانات")
    with tab2:
        st.subheader("تقرير يومي")
        if st.button("عرض التقرير اليومي"):
            today = datetime.now().strftime("%Y-%m-%d")
            daily_expenses = []
            total = 0
            csv_content = load_csv_from_gist()
            if csv_content.strip():
                df = pd.read_csv(io.StringIO(csv_content), encoding="utf-8")
                for _, row in df.iterrows():
                    if row["التاريخ"] == today:
                        daily_expenses.append(row.to_dict())
                        total += float(row["المبلغ"])
            if daily_expenses:
                st.write(f"تقرير اليوم: {today}")
                for expense in daily_expenses:
                    st.write(f"القسم: {expense['القسم']}, المبلغ: {expense['المبلغ']}, ملاحظات: {expense['ملاحظات']}")
                st.write(f"الإجمالي: {total:.2f} جنيه")
            else:
                st.write(f"لا توجد مصروفات لليوم {today}")
    with tab3:
        st.subheader("تقرير شهري")
        if st.button("عرض التقرير الشهري"):
            current_month = datetime.now().strftime("%Y-%m")
            monthly_expenses = []
            total = 0
            csv_content = load_csv_from_gist()
            if csv_content.strip():
                df = pd.read_csv(io.StringIO(csv_content), encoding="utf-8")
                for _, row in df.iterrows():
                    if row["التاريخ"].startswith(current_month):
                        monthly_expenses.append(row.to_dict())
                        total += float(row["المبلغ"])
            if monthly_expenses:
                st.write(f"تقرير الشهر: {current_month}")
                for expense in monthly_expenses:
                    st.write(f"التاريخ: {expense['التاريخ']}, القسم: {expense['القسم']}, المبلغ: {expense['المبلغ']}, ملاحظات: {expense['ملاحظات']}")
                st.write(f"الإجمالي: {total:.2f} جنيه")
            else:
                st.write(f"لا توجد مصروفات للشهر {current_month}")
    with tab4:
        st.subheader("تقرير فيزا وكاش شهري")
        if st.button("عرض تقرير فيزا وكاش الشهري"):
            current_month = datetime.now().strftime("%Y-%m")
            visa_total = 0
            cash_total = 0
            special_total = 0
            expenses_total = 0
            visa_expenses = []
            cash_expenses = []
            special_expenses = []
            expenses_expenses = []
            csv_content = load_csv_from_gist()
            if csv_content.strip():
                df = pd.read_csv(io.StringIO(csv_content), encoding="utf-8")
                for _, row in df.iterrows():
                    if row["التاريخ"].startswith(current_month):
                        if row["القسم"] == "فيزا":
                            visa_total += float(row["المبلغ"])
                            visa_expenses.append(row.to_dict())
                        elif row["القسم"] == "كاش":
                            cash_total += float(row["المبلغ"])
                            cash_expenses.append(row.to_dict())
                        elif row["القسم"] == "مصروفات خاصة":
                            special_total += float(row["المبلغ"])
                            special_expenses.append(row.to_dict())
                        elif row["القسم"] == "مصروفات":
                            expenses_total += float(row["المبلغ"])
                            expenses_expenses.append(row.to_dict())
            if visa_expenses or cash_expenses or special_expenses or expenses_expenses:
                st.write(f"تقرير فيزا وكاش ومصروفات للشهر: {current_month}")
                if visa_expenses:
                    st.write("فيزا:")
                    for expense in visa_expenses:
                        st.write(f"  التاريخ: {expense['التاريخ']}, المبلغ: {expense['المبلغ']}, ملاحظات: {expense['ملاحظات']}")
                    st.write(f"إجمالي الفيزا: {visa_total:.2f} جنيه")
                if cash_expenses:
                    st.write("كاش:")
                    for expense in cash_expenses:
                        st.write(f"  التاريخ: {expense['التاريخ']}, المبلغ: {expense['المبلغ']}, ملاحظات: {expense['ملاحظات']}")
                    st.write(f"إجمالي الكاش: {cash_total:.2f} جنيه")
                if expenses_expenses:
                    st.write("مصروفات:")
                    for expense in expenses_expenses:
                        st.write(f"  التاريخ: {expense['التاريخ']}, المبلغ: {expense['المبلغ']}, ملاحظات: {expense['ملاحظات']}")
                    st.write(f"إجمالي المصروفات: {expenses_total:.2f} جنيه")
                net_visa_cash = visa_total + cash_total - expenses_total
                st.write(f"صافي الفيزا والكاش (بعد خصم المصروفات): {net_visa_cash:.2f} جنيه")
                if special_expenses:
                    st.write("مصروفات خاصة:")
                    for expense in special_expenses:
                        st.write(f"  التاريخ: {expense['التاريخ']}, المبلغ: {expense['المبلغ']}, ملاحظات: {expense['ملاحظات']}")
                    st.write(f"إجمالي المصروفات الخاصة: {special_total:.2f} جنيه")
                total = net_visa_cash + special_total
                st.write(f"الإجمالي الكلي: {total:.2f} جنيه")
            else:
                st.write(f"لا توجد مصروفات فيزا أو كاش أو خاصة أو مصروفات للشهر {current_month}")

def main():
    init_file()
    if MODE == 'desktop':
        root = tk.Tk()
        root.title("إدارة المصروفات Salt&Crunch ")
        root.geometry("1000x800")
        root.configure(bg='#f0f0f0')

        # Set a modern font
        default_font = ('Arial', 12, 'bold')
        title_font = ('Arial', 16, 'bold')

        # Load background image
        original_bg_image = Image.open("WhatsApp Image 2025-12-25 at 12.03.46 AM.jpeg")

        # Create canvas for background
        canvas = tk.Canvas(root, bg='#f0f0f0')
        canvas.pack(fill="both", expand=True)

        # Resize image to fixed 1000x700
        resized_image = original_bg_image.resize((1000, 700), Image.Resampling.BICUBIC)
        bg_photo = ImageTk.PhotoImage(resized_image)
        canvas.create_image(0, 0, image=bg_photo, anchor="nw")
        canvas.image = bg_photo  # Keep a reference to prevent garbage collection

        # Title label
        title_label = tk.Label(root, text="إدارة المصروفات Salt&Crunch", font=title_font, bg='#4a90e2', fg='white', padx=20, pady=10)
        title_label.place(relx=0.5, rely=0.1, anchor="center")

        # Floating button for reports and closure at top left
        closure_button = tk.Button(root, text="تقارير وإغلاق", command=lambda: reports_and_closure_gui(root), bg='#28a745', fg='white', font=('Arial', 10, 'bold'), relief='raised', bd=2, padx=10, pady=5)
        closure_button.place(x=10, y=10)

        frame = tk.Frame(root, bg='#f0f0f0', bd=0, relief='flat')
        frame.place(relx=0.5, rely=0.55, anchor="center")

        button_style = {
            'font': default_font,
            'bg': '#ff8c00',
            'fg': 'white',
            'activebackground': '#ff6600',
            'activeforeground': 'white',
            'relief': 'raised',
            'bd': 2,
            'padx': 20,
            'pady': 10,
            'width': 30,
            'height': 2
        }

        tk.Button(frame, text="إضافة مصروف", command=lambda: add_expense_gui(root), **button_style).pack(side=tk.TOP, pady=10)
        tk.Button(frame, text="إضافة قسم", command=lambda: add_category_gui(root), **button_style).pack(side=tk.TOP, pady=10)
        tk.Button(frame, text="عرض كل المصروفات", command=lambda: show_expenses_gui(root), **button_style).pack(side=tk.TOP, pady=10)
        tk.Button(frame, text="إجمالي المصروفات حسب القسم", command=lambda: total_by_category_gui(root), **button_style).pack(side=tk.TOP, pady=10)
        tk.Button(frame, text="تقارير شهرية", command=lambda: monthly_reports_gui(root), **button_style).pack(side=tk.TOP, pady=10)
        tk.Button(frame, text="تقارير شهرية مفصلة", command=lambda: detailed_monthly_reports_gui(root), **button_style).pack(side=tk.TOP, pady=10)
        tk.Button(frame, text="خروج", command=root.quit, bg='#e74c3c', activebackground='#c0392b', **{k: v for k, v in button_style.items() if k not in ['bg', 'activebackground']}).pack(side=tk.TOP, pady=10)

        root.mainloop()
    else:
        st.set_page_config(page_title="إدارة المصروفات Salt&Crunch", page_icon="💰", layout="wide")
        st.title("إدارة المصروفات Salt&Crunch")
        # Load and display background image
        try:
            bg_image = Image.open("WhatsApp Image 2025-12-25 at 12.03.46 AM.jpeg")
            st.image(bg_image, use_column_width=True)
        except FileNotFoundError:
            pass
        # Sidebar for navigation
        st.sidebar.title("القائمة")
        page = st.sidebar.radio("اختر الصفحة", ["إضافة مصروف", "إضافة قسم", "عرض كل المصروفات", "إجمالي المصروفات حسب القسم", "تقارير شهرية", "تقارير شهرية مفصلة", "تقارير وإغلاق", "مزامنة البيانات"])
        if page == "إضافة مصروف":
            add_expense_web()
        elif page == "إضافة قسم":
            add_category_web()
        elif page == "عرض كل المصروفات":
            show_expenses_web()
        elif page == "إجمالي المصروفات حسب القسم":
            total_by_category_web()
        elif page == "تقارير شهرية":
            monthly_reports_web()
        elif page == "تقارير شهرية مفصلة":
            detailed_monthly_reports_web()
        elif page == "تقارير وإغلاق":
            reports_and_closure_web()
        elif page == "مزامنة البيانات":
            sync_data_web()

if __name__ == "__main__":
    main()
